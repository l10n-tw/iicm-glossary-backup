import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def main():
    csv_dir = Path("iicm-glossary-csv")
    output_file = "iicm_glossary.xlsx"
    
    # Get all CSV files and sort them alphabetically
    csv_files = sorted(csv_dir.glob("termb_*.csv"))
    
    if not csv_files:
        print("No CSV files found!")
        return
    
    print(f"Found {len(csv_files)} CSV files")
    
    # Create workbook
    wb = Workbook()
    # Remove default worksheet
    if wb.active:
        wb.remove(wb.active)
    
    for csv_file in csv_files:
        # Read CSV file
        print(f"Processing: {csv_file.name}")
        
        # Create worksheet
        sheet_name = csv_file.stem.replace('termb_', '')
        ws = wb.create_sheet(title=sheet_name)
        
        # Read CSV and write to worksheet
        with open(csv_file, 'r', encoding='utf-8') as f:
            csv_reader = csv.reader(f)
            for row_idx, row in enumerate(csv_reader, start=1):
                for col_idx, value in enumerate(row, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Auto-adjust column width
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Maximum width limited to 50
            ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save Excel file
    wb.save(output_file)
    print(f"\nCompleted! Excel file saved as: {output_file}")
    print(f"Created {len(csv_files)} worksheets")

if __name__ == "__main__":
    main()
